```python
"""
Adaptive Valuation Screener — Streamlit App
=============================================
Webinterface voor de Nasdaq-100 / Euro Stoxx 50 valuation screener.

LOKAAL DRAAIEN:
    pip install -r requirements.txt
    streamlit run app.py

GRATIS HOSTEN (Streamlit Community Cloud):
    1. Zet app.py + requirements.txt in een publieke GitHub-repo
    2. Ga naar https://share.streamlit.io -> "New app"
    3. Kies je repo, branch, en app.py als entry point -> Deploy
    4. Je krijgt een gratis URL zoals https://jouwnaam-screener.streamlit.app
"""

import re
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

st.set_page_config(page_title="Adaptive Valuation Screener", layout="wide")

# =============================================================================
# CONFIG
# =============================================================================

@dataclass
class ScreenerConfig:
    band_mult: float
    inner_ratio: float
    mid_ratio: float
    outer_ext: float
    percentile_lookback: int
    cycle_roc_length: int
    base_dca: float
    dca_pct_very_cheap: float
    dca_pct_cheap: float
    dca_pct_fair: float
    dca_pct_expensive: float
    dca_pct_very_expensive: float


ZONE_COLORS = {
    "Very Cheap":     "2962FF",
    "Cheap":          "00BCD4",
    "Fair Value":     "4CAF50",
    "Expensive":      "FF9800",
    "Very Expensive": "F23645",
}

NASDAQ100_URLS = [
    "https://en.wikipedia.org/wiki/List_of_NASDAQ-100_companies",
    "https://en.wikipedia.org/wiki/Nasdaq-100",
]

NASDAQ100_BACKUP_CORE = [
    "AAPL", "MSFT", "NVDA", "AMZN", "GOOGL", "GOOG", "META", "AVGO", "TSLA",
    "COST", "NFLX", "ASML", "AMD", "PEP", "ADBE", "CSCO", "TMUS", "INTC",
    "INTU", "QCOM", "TXN", "AMGN", "HON", "AMAT", "BKNG", "SBUX", "GILD",
    "ADP", "MDLZ", "ISRG", "VRTX", "REGN", "PANW", "LRCX", "KLAC", "SNPS",
    "CDNS", "MU", "ORLY", "MAR", "PYPL", "CTAS", "ADI", "PCAR", "CRWD",
]

EUROSTOXX50_BACKUP = [
    "ASML.AS", "MC.PA", "TTE.PA", "SAP.DE", "SIE.DE", "OR.PA", "ALV.DE",
    "SU.PA", "SAN.PA", "AIR.PA", "AI.PA", "BNP.PA", "DG.PA", "CS.PA",
    "MBG.DE", "MUV2.DE", "BAS.DE", "BAYN.DE", "DTE.DE", "IFX.DE", "ADS.DE",
    "VOW3.DE", "DB1.DE", "VNA.DE", "DHL.DE", "ENGI.PA", "ACA.PA", "SAF.PA",
    "EL.PA", "STLAM.MI", "ENEL.MI", "ISP.MI", "ENI.MI", "UCG.MI", "RACE.MI",
    "INGA.AS", "ADYEN.AS", "PRX.AS", "AD.AS", "WKL.AS", "PHIA.AS", "SAN.MC",
    "BBVA.MC", "IBE.MC", "ITX.MC", "ABI.BR", "NDA-FI.HE", "KER.PA", "RI.PA",
    "AMS.MC",
]


# =============================================================================
# SYMBOOLLIJSTEN
# =============================================================================

def _extract_tickers_from_table(t: pd.DataFrame) -> list[str]:
    ticker_cols = [c for c in t.columns if "ticker" in str(c).strip().lower()
                   or str(c).strip().lower() == "symbol"]
    if not ticker_cols:
        return []
    raw = t[ticker_cols[0]].astype(str).tolist()
    out = []
    for val in raw:
        for sym in re.split(r"[,/]", val):
            sym = re.sub(r"\[.*?\]", "", sym).strip().upper()
            if sym and sym != "NAN" and re.match(r"^[A-Z]{1,5}([.\-][A-Z])?$", sym):
                out.append(sym.replace(".", "-"))
    return out


@st.cache_data(ttl=86400, show_spinner=False)
def get_nasdaq100_tickers() -> tuple[list[str], str]:
    for url in NASDAQ100_URLS:
        try:
            tables = pd.read_html(url)
        except Exception:
            continue
        best: list[str] = []
        for t in tables:
            found = _extract_tickers_from_table(t)
            if len(found) > len(best):
                best = found
        tickers = sorted(set(best))
        if len(tickers) >= 90:
            return tickers, f"live ({url})"
    return NASDAQ100_BACKUP_CORE, "back-up (kernlijst, niet volledig)"


@st.cache_data(ttl=86400, show_spinner=False)
def get_eurostoxx50_tickers() -> tuple[list[str], str]:
    url = "https://en.wikipedia.org/wiki/EURO_STOXX_50"
    try:
        tables = pd.read_html(url)
        for t in tables:
            ticker_cols = [c for c in t.columns if any(
                key in str(c).strip().lower() for key in ("ticker", "symbol")
            )]
            if ticker_cols and 40 <= len(t) <= 55:
                raw = t[ticker_cols[0]].astype(str).str.strip().tolist()
                raw = [x for x in raw if x and x.lower() != "nan"]
                if len(raw) >= 40:
                    return sorted(set(raw)), "live (Wikipedia)"
    except Exception:
        pass
    return EUROSTOXX50_BACKUP, "back-up lijst"


# =============================================================================
# REGRESSIEMODELLEN
# =============================================================================

def _ols(x: np.ndarray, y: np.ndarray):
    n = len(x)
    sx, sy = x.sum(), y.sum()
    sxy, sx2 = (x * y).sum(), (x * x).sum()
    denom = n * sx2 - sx * sx
    if denom == 0:
        return np.nan, np.nan, np.nan, np.nan
    b = (n * sxy - sx * sy) / denom
    a = (sy - b * sx) / n
    pred = a + b * x
    ss_res = np.sum((y - pred) ** 2)
    ss_tot = np.sum((y - y.mean()) ** 2)
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else np.nan
    std_res = np.sqrt(max(ss_res, 0) / max(n - 2, 1))
    return a, b, r2, std_res


def fit_best_model(prices: np.ndarray):
    n = len(prices)
    x = np.arange(n, dtype=float)
    xlog = np.log(x + 1.0)
    y = prices.astype(float)
    ylog = np.log(y)
    candidates = {
        "Linear":      dict(zip(("a", "b", "r2", "std"), _ols(x, y)),       use_log_x=False, mult=False),
        "Exponential": dict(zip(("a", "b", "r2", "std"), _ols(x, ylog)),    use_log_x=False, mult=True),
        "Logarithmic": dict(zip(("a", "b", "r2", "std"), _ols(xlog, y)),    use_log_x=True,  mult=False),
        "Power Law":   dict(zip(("a", "b", "r2", "std"), _ols(xlog, ylog)), use_log_x=True,  mult=True),
    }
    best_name, best = None, None
    for name, c in candidates.items():
        r2 = c["r2"]
        if r2 is None or np.isnan(r2):
            continue
        if best is None or r2 > best["r2"]:
            best_name, best = name, c
    if best is None:
        raise ValueError("Kon geen enkel regressiemodel fitten.")
    best["name"] = best_name
    return best


def predict(model: dict, idx: float) -> float:
    xlog = np.log(idx + 1.0)
    raw = model["a"] + model["b"] * (xlog if model["use_log_x"] else idx)
    return np.exp(raw) if model["mult"] else raw


def band(model: dict, base: float, k: float) -> float:
    return base * np.exp(k * model["std"]) if model["mult"] else base + k * model["std"]


def deviation_series(model: dict, prices: np.ndarray) -> np.ndarray:
    n = len(prices)
    idx = np.arange(n, dtype=float)
    pred = np.array([predict(model, i) for i in idx])
    if model["mult"]:
        return np.log(prices / pred) / model["std"]
    return (prices - pred) / model["std"]


def compute_metrics(symbol: str, prices: pd.Series, cfg: ScreenerConfig, index_name: str) -> dict:
    prices = prices.dropna()
    n = len(prices)
    if n < 30:
        return {"Symbol": symbol, "Index": index_name, "Error": f"Te weinig data ({n} bars)"}

    y = prices.to_numpy()
    model = fit_best_model(y)

    k1 = cfg.band_mult * cfg.inner_ratio
    k2 = cfg.band_mult * (cfg.inner_ratio + cfg.mid_ratio)
    k3 = cfg.band_mult * (cfg.inner_ratio + cfg.mid_ratio + cfg.outer_ext)

    last_idx = n - 1
    reg_value = predict(model, last_idx)
    src_price = float(y[-1])

    B1 = band(model, reg_value, -k2)
    B2 = band(model, reg_value, -k1)
    B3 = band(model, reg_value, k1)
    B4 = band(model, reg_value, k2)

    dev_series = deviation_series(model, y)
    deviation = float(dev_series[-1])

    dca_score = float(np.clip(50 - (deviation / k3) * 50, 0, 100))
    fv_discount = (src_price - reg_value) / reg_value * 100

    lb = min(cfg.percentile_lookback, n)
    hist_window = dev_series[-lb:]
    percent_rank = float((hist_window < deviation).sum() / len(hist_window) * 100)
    cheaper_than_pct = 100 - percent_rank

    if src_price < B1:
        zone, zone_dca_pct = "Very Cheap", cfg.dca_pct_very_cheap
    elif src_price < B2:
        zone, zone_dca_pct = "Cheap", cfg.dca_pct_cheap
    elif src_price < B3:
        zone, zone_dca_pct = "Fair Value", cfg.dca_pct_fair
    elif src_price < B4:
        zone, zone_dca_pct = "Expensive", cfg.dca_pct_expensive
    else:
        zone, zone_dca_pct = "Very Expensive", cfg.dca_pct_very_expensive

    suggested_dca = cfg.base_dca * zone_dca_pct / 100

    conf_r2 = max(0.0, model["r2"]) * 100 if not np.isnan(model["r2"]) else 0.0
    conf_bars = min(n / 2000, 1.0) * 100
    confidence = round(conf_r2 * 0.7 + conf_bars * 0.3)

    roc_len = min(cfg.cycle_roc_length, n - 1)
    price_roc = (src_price - y[-1 - roc_len]) / y[-1 - roc_len] * 100 if roc_len > 0 else 0.0

    if deviation < -k2 and price_roc < 0:
        cycle = "Accumulation Phase"
    elif deviation < -k1:
        cycle = "Early Bull Phase"
    elif deviation < k1 and price_roc > 0:
        cycle = "Bull Phase"
    elif deviation < k2 and price_roc > 0:
        cycle = "Late Bull Phase"
    elif deviation >= k2:
        cycle = "Distribution Phase"
    elif price_roc < 0:
        cycle = "Bear Phase"
    else:
        cycle = "Neutral Phase"

    return {
        "Symbol": symbol, "Index": index_name,
        "Price": round(src_price, 4), "Fair Value": round(reg_value, 4),
        "Discount %": round(fv_discount, 2), "Zone": zone,
        "DCA Score": round(dca_score, 1), "Percentile %": round(percent_rank, 1),
        "Cheaper Than %": round(cheaper_than_pct, 1), "Suggested DCA": round(suggested_dca, 2),
        "DCA %": zone_dca_pct, "Confidence %": confidence, "Cycle": cycle,
        "Model": model["name"], "R2": round(model["r2"], 4) if not np.isnan(model["r2"]) else None,
        "Bars": n,
    }


def fetch_prices(symbols, period, interval, batch_size, progress_cb=None):
    out = {}
    n_batches = (len(symbols) + batch_size - 1) // batch_size
    for i in range(0, len(symbols), batch_size):
        batch = symbols[i:i + batch_size]
        batch_num = i // batch_size + 1
        try:
            data = yf.download(batch, period=period, interval=interval, progress=False,
                                auto_adjust=True, group_by="ticker", threads=True)
        except Exception:
            data = None
        if data is not None:
            for sym in batch:
                try:
                    s = data["Close"] if len(batch) == 1 else data[sym]["Close"]
                    s = s.dropna()
                    if len(s) > 0:
                        out[sym] = s
                except Exception:
                    pass
        if progress_cb:
            progress_cb(batch_num, n_batches)
        if batch_num < n_batches:
            time.sleep(1)
    return out


def build_excel_bytes(rows, cfg) -> bytes:
    valid = [r for r in rows if "Error" not in r]
    errors = [r for r in rows if "Error" in r]
    valid.sort(key=lambda r: r["DCA Score"], reverse=True)

    columns = ["Symbol", "Index", "Price", "Fair Value", "Discount %", "Zone", "DCA Score",
               "Percentile %", "Cheaper Than %", "Suggested DCA", "DCA %",
               "Confidence %", "Cycle", "Model", "R2", "Bars"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Screener"
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    body_font = Font(name="Arial", size=10)

    for col_idx, name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(valid, start=2):
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(name))
            cell.font = body_font
            if name == "Zone":
                color = ZONE_COLORS.get(r.get("Zone"), "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)

    dca_col = columns.index("DCA Score") + 1
    dca_letter = get_column_letter(dca_col)
    last_row = max(len(valid) + 1, 2)
    ws.conditional_formatting.add(f"{dca_letter}2:{dca_letter}{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(f"{dca_letter}2:{dca_letter}{last_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["30"], fill=PatternFill("solid", fgColor="FFC7CE")))
    for col_idx, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(name) + 2)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"

    if errors:
        ws2 = wb.create_sheet("Skipped")
        ws2.cell(row=1, column=1, value="Symbol").font = header_font
        ws2.cell(row=1, column=2, value="Index").font = header_font
        ws2.cell(row=1, column=3, value="Reden").font = header_font
        for i, r in enumerate(errors, start=2):
            ws2.cell(row=i, column=1, value=r["Symbol"])
            ws2.cell(row=i, column=2, value=r.get("Index", ""))
            ws2.cell(row=i, column=3, value=r["Error"])

    ws3 = wb.create_sheet("Config")
    ws3.cell(row=1, column=1, value="Parameter").font = header_font
    ws3.cell(row=1, column=2, value="Waarde").font = header_font
    for i, (k, v) in enumerate(asdict(cfg).items(), start=2):
        ws3.cell(row=i, column=1, value=k)
        ws3.cell(row=i, column=2, value=v)
    ws3.cell(row=len(asdict(cfg)) + 3, column=1, value="Gegenereerd op")
    ws3.cell(row=len(asdict(cfg)) + 3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# STREAMLIT UI
# =============================================================================

st.title("📊 Adaptive Valuation Screener")
st.caption("Nasdaq-100 + Euro Stoxx 50 — regressie-gebaseerde waardering & DCA-score")

with st.sidebar:
    st.header("Universum")
    use_nasdaq = st.checkbox("Nasdaq-100", value=True)
    use_stoxx = st.checkbox("Euro Stoxx 50", value=True)
    custom_extra = st.text_input("Extra tickers (komma-gescheiden, optioneel)", "")

    st.header("Data")
    period = st.selectbox("Periode", ["1y", "2y", "5y", "10y", "max"], index=2)
    interval = st.selectbox("Interval", ["1d", "1wk"], index=0)
    batch_size = st.slider("Batchgrootte (download)", 5, 40, 20)

    st.header("Banden")
    band_mult = st.slider("Band Multiplier (σ)", 0.5, 4.0, 2.0, 0.1)
    inner_ratio = st.slider("Fair Value zone (× σ)", 0.05, 1.0, 0.35, 0.05)
    mid_ratio = st.slider("Cheap/Expensive zone (× σ)", 0.1, 2.0, 1.0, 0.05)
    outer_ext = st.slider("Very Cheap/Expensive extensie (× σ)", 0.1, 2.0, 1.0, 0.05)

    st.header("DCA Allocation")
    base_dca = st.number_input("Basisbedrag", value=100.0, step=10.0)
    dca_vc = st.number_input("Very Cheap %", value=300.0, step=10.0)
    dca_c = st.number_input("Cheap %", value=150.0, step=10.0)
    dca_f = st.number_input("Fair Value %", value=100.0, step=10.0)
    dca_e = st.number_input("Expensive %", value=50.0, step=10.0)
    dca_ve = st.number_input("Very Expensive %", value=0.0, step=10.0)

    st.header("Overig")
    percentile_lookback = st.number_input("Percentiel lookback (bars)", value=2000, step=100)
    cycle_roc_length = st.number_input("Cycle momentum lookback (bars)", value=30, step=5)

    run_btn = st.button("🔍 Screener uitvoeren", type="primary", use_container_width=True)

if run_btn:
    symbol_index_map: dict[str, str] = {}

    if use_nasdaq:
        with st.spinner("Nasdaq-100 lijst ophalen..."):
            nq, src = get_nasdaq100_tickers()
        st.info(f"Nasdaq-100: {len(nq)} tickers ({src})")
        symbol_index_map.update({s: "Nasdaq-100" for s in nq})

    if use_stoxx:
        with st.spinner("Euro Stoxx 50 lijst ophalen..."):
            sx, src = get_eurostoxx50_tickers()
        st.info(f"Euro Stoxx 50: {len(sx)} tickers ({src})")
        symbol_index_map.update({s: "Euro Stoxx 50" for s in sx})

    if custom_extra.strip():
        extra = [s.strip().upper() for s in custom_extra.split(",") if s.strip()]
        symbol_index_map.update({s: "Custom" for s in extra})

    all_symbols = list(symbol_index_map.keys())
    if not all_symbols:
        st.error("Selecteer minstens één universum of geef eigen tickers op.")
        st.stop()

    cfg = ScreenerConfig(
        band_mult=band_mult, inner_ratio=inner_ratio, mid_ratio=mid_ratio, outer_ext=outer_ext,
        percentile_lookback=int(percentile_lookback), cycle_roc_length=int(cycle_roc_length),
        base_dca=base_dca, dca_pct_very_cheap=dca_vc, dca_pct_cheap=dca_c, dca_pct_fair=dca_f,
        dca_pct_expensive=dca_e, dca_pct_very_expensive=dca_ve,
    )

    progress_bar = st.progress(0.0, text="Data downloaden...")

    def on_progress(batch_num, n_batches):
        progress_bar.progress(batch_num / n_batches, text=f"Data downloaden... batch {batch_num}/{n_batches}")

    price_data = fetch_prices(all_symbols, period, interval, batch_size, on_progress)

    rows = []
    calc_bar = st.progress(0.0, text="Metrics berekenen...")
    for i, sym in enumerate(all_symbols, start=1):
        idx_name = symbol_index_map[sym]
        if sym not in price_data:
            rows.append({"Symbol": sym, "Index": idx_name, "Error": "Geen data opgehaald"})
        else:
            try:
                rows.append(compute_metrics(sym, price_data[sym], cfg, idx_name))
            except Exception as e:
                rows.append({"Symbol": sym, "Index": idx_name, "Error": str(e)})
        calc_bar.progress(i / len(all_symbols), text=f"Metrics berekenen... {i}/{len(all_symbols)}")

    progress_bar.empty()
    calc_bar.empty()

    valid_rows = [r for r in rows if "Error" not in r]
    error_rows = [r for r in rows if "Error" in r]

    st.success(f"Klaar — {len(valid_rows)}/{len(all_symbols)} symbolen verwerkt.")

    if valid_rows:
        df = pd.DataFrame(valid_rows).sort_values("DCA Score", ascending=False).reset_index(drop=True)

        def zone_color(val):
            c = ZONE_COLORS.get(val, "FFFFFF")
            return f"background-color: #{c}; color: white; font-weight: bold;"

        styled_df = (
    df.style
      .map(zone_color, subset=["Zone"])
      .background_gradient(
            subset=["DCA Score"],
            cmap="RdYlGn",
            vmin=0,
            vmax=100
      )
)

        st.dataframe(
            styled_df,
            width="stretch",
            height=600,
        )

        excel_bytes = build_excel_bytes(rows, cfg)
        st.download_button(
            "⬇️ Download als Excel",
            data=excel_bytes,
            file_name=f"valuation_screener_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if error_rows:
        with st.expander(f"⚠️ {len(error_rows)} overgeslagen symbolen"):
            st.dataframe(pd.DataFrame(error_rows), use_container_width=True)
else:
    st.info("Stel je parameters in de zijbalk in en klik op **Screener uitvoeren**.")
```